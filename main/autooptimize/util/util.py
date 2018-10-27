import ctypes, logging.config, os, platform, re, shutil, sys, time, win32gui
from configparser import ConfigParser
from ctypes import wintypes
from random import randint
import psutil, requests, win32.win32api as win32api, win32.win32evtlog as win32evtlog, win32con
from lxml import etree
from autooptimize.globalEnvStorage import GlobalEnvStorage

def get_free_space_mb(folder):
    """ Return folder/drive free space (in bytes)
    """
    try:
        if platform.system() == 'Windows':
            free_bytes = ctypes.c_ulonglong(0)
            ctypes.windll.kernel32.GetDiskFreeSpaceExW(ctypes.c_wchar_p(folder), None, None, ctypes.pointer(free_bytes))
            return free_bytes.value / 1024 / 1024
        st = os.statvfs(folder)
        return st.f_bavail * st.f_frsize / 1024
    except:
        return 1000


def getCity():
    try:
        re = requests.get('http://ws.sj.qq.com/webservices/ipAddress!getIpAndTime.do', timeout=10)
        GlobalEnvStorage.infoLogger.info('%s', re.text.split('|')[0])
        re = requests.get('http://ip.taobao.com/service/getIpInfo.php?ip=' + re.text.split('|')[0], timeout=10)
        city = re.json()['data']['city']
        GlobalEnvStorage.infoLogger.info('%s', city)
        return city
    except:
        return '---'


def isOnline():
    result = os.system('ping 8.8.8.8')
    if result:
        GlobalEnvStorage.infoLogger.info('ping fail,没有网咯')
    else:
        GlobalEnvStorage.infoLogger.info('ping ok')
    return result


def connect():
    while True:
        name = GlobalEnvStorage.connectName
        username = GlobalEnvStorage.connectUserName
        password = GlobalEnvStorage.connectPassword
        cmd_str = 'rasdial %s %s %s' % (name, username, password)
        res = os.system(cmd_str)
        GlobalEnvStorage.infoLogger.info('拨号命令退出')
        if res == 0:
            GlobalEnvStorage.infoLogger.info('connect successful')
            break
        else:
            GlobalEnvStorage.infoLogger.info('%s', res)
            GlobalEnvStorage.infoLogger.info('出现错误继续拨号')
            time.sleep(10)


def disConnect():
    name = GlobalEnvStorage.connectName
    cmdstr = 'rasdial %s /disconnect' % name
    os.system(cmdstr)


def reConnect():
    disConnect()
    time.sleep(2)
    connect()


def initResolution():
    mode = win32api.EnumDisplaySettings()
    mode.PelsWidth = int(GlobalEnvStorage.PelsWidth)
    mode.PelsHeight = int(GlobalEnvStorage.PelsHeight)
    mode.BitsPerPel = int(GlobalEnvStorage.BitsPerPel)
    win32api.ChangeDisplaySettings(mode, 0)


def changeCustomerKeyword(customerKeyword):
    # GlobalEnvStorage.infoLogger.info('%s\n', customerKeyword)
    customerKeyword.kuaizhaoPercent = valueToPercentage(customerKeyword.kuaizhaoPercent)
    customerKeyword.baiduSemPercent = valueToPercentage(customerKeyword.baiduSemPercent)
    customerKeyword.dragPercent = valueToPercentage(customerKeyword.dragPercent)
    customerKeyword.zhanneiPercent = valueToPercentage(customerKeyword.zhanneiPercent)
    customerKeyword.zhanwaiPercent = valueToPercentage(customerKeyword.zhanwaiPercent)
    if customerKeyword.pageSize == 0:
        customerKeyword.pageSize = 10
    elif customerKeyword.pageSize == 1:
        customerKeyword.pageSize = 20
    else:
        customerKeyword.pageSize = 50
    customerKeyword.pageRemainMinTime = customerKeyword.pageRemainMinTime / 1000
    customerKeyword.pageRemainMaxTime = customerKeyword.pageRemainMaxTime / 1000
    customerKeyword.inputDelayMinTime = customerKeyword.inputDelayMinTime / 1000
    customerKeyword.inputDelayMaxTime = customerKeyword.inputDelayMaxTime / 1000
    customerKeyword.slideDelayMinTime = customerKeyword.slideDelayMinTime / 1000
    customerKeyword.slideDelayMaxTime = customerKeyword.slideDelayMaxTime / 1000
    customerKeyword.waitTimeAfterOpenBaidu = customerKeyword.waitTimeAfterOpenBaidu / 1000
    customerKeyword.waitTimeBeforeClick = customerKeyword.waitTimeBeforeClick / 1000
    customerKeyword.waitTimeAfterClick = customerKeyword.waitTimeAfterClick / 1000
    customerKeyword.titleRemainMinTime = customerKeyword.titleRemainMinTime / 1000
    customerKeyword.titleRemainMaxTime = customerKeyword.titleRemainMaxTime / 1000
    return customerKeyword


def valueToPercentage(val):
    if val == 0:
        percentage = 0
    else:
        if val == 1:
            percentage = 10
        else:
            if val == 2:
                percentage = 30
            else:
                if val == 3:
                    percentage = 50
                else:
                    percentage = 100
    return percentage


def UAconf():
    PC = []
    Phone = []
    PC_useful_lv1 = []
    PC_useful_lv2 = []
    PC_useful_lv3 = []
    PC_useful_lv4 = []
    Phone_useful_lv1 = []
    Phone_useful_lv2 = []
    Phone_useful_lv3 = []
    Phone_useful_lv4 = []
    filePath = 'C:\\Users\\hp\\Desktop\\ua.xlsx'
    import openpyxl
    wb = openpyxl.load_workbook(filename=filePath)
    sheetnames = wb.get_sheet_names()
    sheet = wb.get_sheet_by_name(sheetnames[0])
    for rowNum in range(2, sheet.max_row):
        try:
            ua = sheet.cell(row=rowNum, column=1).value
            pixelRatio = float(sheet.cell(row=rowNum, column=4).value)
            height = int(sheet.cell(row=rowNum, column=3).value)
            width = int(sheet.cell(row=rowNum, column=2).value)
            count = int(sheet.cell(row=rowNum, column=5).value)
            if ua != None:
                pass
            if 'spider' not in ua and 'Download' not in ua and 'MSIE' not in ua:
                if 'Windows' in ua and 'Phone' not in ua:
                    a = {'userAgent': ua,  'deviceMetrics': {'pixelRatio': pixelRatio,  'height': height,  'width': width},  'count': count}
                    PC.append(a)
                if 'Android' in ua or 'iPhone' in ua:
                    if 'Android 2' not in ua and 'Android 3' not in ua and 'Android 4' not in ua and 'Android/4' not in ua:
                        if height != 0 and width != 0 and height <= 640 and height >= 300 and width < height:
                            if 'baiduboxapp' not in ua:
                                a = {'userAgent': ua,  'deviceMetrics': {'pixelRatio': pixelRatio,  'height': height,  'width': width},  'count': count}
                                Phone.append(a)
        except:
            pass

    wb.close()
    print(PC.__str__())
    print(PC.__len__())
    print(Phone.__str__())
    print(Phone.__len__())
    cf = ConfigParser()
    cf.read('..//..//config//dev//UAList.conf', encoding='utf-8-sig')
    for idx in range(len(PC)):
        try:
            cf.set('uatry', 'PC_' + str(idx), str(PC[idx]))
            if PC[idx]['count'] >= 1000:
                PC_useful_lv1.append(PC[idx])
            else:
                if PC[idx]['count'] >= 100:
                    PC_useful_lv2.append(PC[idx])
                else:
                    if PC[idx]['count'] >= 10:
                        PC_useful_lv3.append(PC[idx])
                    else:
                        PC_useful_lv4.append(PC[idx])
        except:
            print('大概是乱码' + str(PC[idx]))

    for idx in range(len(Phone)):
        try:
            cf.set('uatry', 'Phone_' + str(idx), str(Phone[idx]))
            if Phone[idx]['count'] >= 1000:
                Phone_useful_lv1.append(Phone[idx])
            else:
                if Phone[idx]['count'] >= 100:
                    Phone_useful_lv2.append(Phone[idx])
                else:
                    if Phone[idx]['count'] >= 10:
                        Phone_useful_lv3.append(Phone[idx])
                    else:
                        Phone_useful_lv4.append(Phone[idx])
        except:
            print('大概是乱码' + str(Phone[idx]))

    cf.set('PCList', 'PC_lv1', str(PC_useful_lv1))
    cf.set('PCList', 'PC_lv2', str(PC_useful_lv2))
    cf.set('PCList', 'PC_lv3', str(PC_useful_lv3))
    cf.set('PCList', 'PC_lv4', str(PC_useful_lv4))
    cf.set('PhoneList', 'Phone_lv1', str(Phone_useful_lv1))
    cf.set('PhoneList', 'Phone_lv2', str(Phone_useful_lv2))
    cf.set('PhoneList', 'Phone_lv3', str(Phone_useful_lv3))
    cf.set('PhoneList', 'Phone_lv4', str(Phone_useful_lv4))
    cf.write(open('..//..//config//dev//UAList.conf', 'w', encoding='utf-8-sig'))


def getUA():
    cf = ConfigParser()
    if GlobalEnvStorage.env != 'Development':
        cf.read('..//config//dev//UAList.conf', encoding='utf-8-sig')
    else:
        cf.read('C:\\working' + '\\UAList.conf', encoding='utf-8-sig')
    n = randint(1, 100)
    if GlobalEnvStorage.customerKeyword.terminalType == 'PC':
        if n <= 10:
            PCList = eval(cf.get('PCList', 'pc_lv1'))
        elif n <= 55:
            PCList = eval(cf.get('PCList', 'pc_lv2'))
        elif n <= 90:
            PCList = eval(cf.get('PCList', 'pc_lv3'))
        elif n <= 100:
            PCList = eval(cf.get('PCList', 'pc_lv4'))
        UA = PCList[randint(0, len(PCList) - 1)]
    else:
        if n <= 5:
            PhoneList = eval(cf.get('PhoneList', 'phone_lv1'))
        elif n <= 45:
            PhoneList = eval(cf.get('PhoneList', 'phone_lv2'))
        elif n <= 85:
            PhoneList = eval(cf.get('PhoneList', 'phone_lv3'))
        elif n <= 100:
            PhoneList = eval(cf.get('PhoneList', 'phone_lv4'))
        UA = PhoneList[randint(0, len(PhoneList) - 1)]
    return UA


def clearCookie():
    if GlobalEnvStorage.customerKeyword.clearCookie == 1:
        GlobalEnvStorage.infoLogger.info('每次都清除cookies')
        GlobalEnvStorage.browser.cookies.delete()
    else:
        if GlobalEnvStorage.customerKeyword.clearCookie == 2:
            GlobalEnvStorage.infoLogger.info('随机清除cookies')
            percentage = randint(1, 100)
            if percentage <= GlobalEnvStorage.ClearCookies_clearPercentage:
                GlobalEnvStorage.browser.cookies.delete()
                GlobalEnvStorage.infoLogger.info('随机清除cookies---')
        else:
            if GlobalEnvStorage.customerKeyword.clearCookie == 3:
                GlobalEnvStorage.infoLogger.info('达到次数清除cookies')
                if GlobalEnvStorage.profileIDCountList[str(GlobalEnvStorage.profileID)]['count'] % GlobalEnvStorage.ClearCookies_clearNum == 0:
                    GlobalEnvStorage.browser.cookies.delete()
                    GlobalEnvStorage.infoLogger.info('达到次数清除cookies---')
            else:
                GlobalEnvStorage.infoLogger.info('不清除cookies')


def fetchSuggestionsValue(cunstomerkeyword, terminalType):
    GlobalEnvStorage.infoLogger.info('开始获得下拉相关的词')
    str = None
    try:
        if terminalType == 'PC':
            if GlobalEnvStorage.customerKeyword.searchEngine == '百度':
                url = 'http://suggestion.baidu.com/su'
                fetchSuggestion = {'wd': cunstomerkeyword,  'ie': 'utf-8'}
                headers = {'user-agent': GlobalEnvStorage.UA}
                r = requests.get(url, params=fetchSuggestion, timeout=10, headers=headers)
                text = r.text
                r.close()
                firstPosition = text.index('[')
                lastPositon = text.index(']')
                arrStr = text[firstPosition + 1:lastPositon]
                arr = arrStr.split(',')
                str = arr[randint(0, len(arr) - 1)].replace('"', '')
            else:
                if GlobalEnvStorage.customerKeyword.searchEngine == '搜狗':
                    url = 'https://www.sogou.com/suggnew/ajajjson'
                    fetchSuggestion = {'key': cunstomerkeyword,  'ie': 'utf-8'}
                    headers = {'user-agent': GlobalEnvStorage.UA}
                    r = requests.get(url, params=fetchSuggestion, timeout=10, headers=headers)
                    text = r.text
                    r.close()
                    firstPosition = text.index('[')
                    firstPosition = text.index('[', firstPosition + 1)
                    lastPositon = text.index(']')
                    arrStr = text[firstPosition + 1:lastPositon]
                    arr = arrStr.split(',')
                    str = arr[randint(0, len(arr) - 1)].replace('"', '')
                else:
                    if GlobalEnvStorage.customerKeyword.searchEngine == '360':
                        url = 'https://sug.so.360.cn/suggest'
                        cunstomerkeyword = cunstomerkeyword.encode('GBK')
                        fetchSuggestion = {'word': cunstomerkeyword}
                        headers = {'user-agent': GlobalEnvStorage.UA}
                        r = requests.get(url, params=fetchSuggestion, timeout=10, headers=headers)
                        text = r.text
                        r.close()
                        firstPosition = text.index('[')
                        lastPositon = text.index(']')
                        arrStr = text[firstPosition + 1:lastPositon]
                        arr = arrStr.split(',')
                        str = arr[randint(0, len(arr) - 1)].replace('"', '')
        else:
            if GlobalEnvStorage.customerKeyword.searchEngine == '百度':
                url = 'http://m.baidu.com/su'
                fetchSuggestion = {'wd': cunstomerkeyword,  'ie': 'utf-8'}
                headers = {'user-agent': GlobalEnvStorage.UA}
                r = requests.get(url, params=fetchSuggestion, timeout=10, headers=headers)
                text = r.text
                r.close()
                firstPosition = text.index('[')
                lastPositon = text.index(']')
                arrStr = text[firstPosition + 1:lastPositon]
                arr = arrStr.split(',')
                str = arr[randint(0, len(arr) - 1)].replace('"', '')
            else:
                if GlobalEnvStorage.customerKeyword.searchEngine == '搜狗':
                    url = 'https://m.sogou.com/web/sugg.jsp'
                    fetchSuggestion = {'kw': cunstomerkeyword,  'ie': 'utf-8'}
                    headers = {'user-agent': GlobalEnvStorage.UA}
                    r = requests.get(url, params=fetchSuggestion, timeout=10, headers=headers)
                    text = r.text
                    r.close()
                    firstPosition = text.index('[')
                    lastPositon = text.index(']')
                    arrStr = text[firstPosition + 1:lastPositon]
                    arr = arrStr.split(',')
                    str = arr[randint(0, len(arr) - 1)].replace('"', '')
                else:
                    if GlobalEnvStorage.customerKeyword.searchEngine == '神马':
                        url = 'http://sugs.m.sm.cn/web'
                        fetchSuggestion = {'t': 'w',  'q': cunstomerkeyword}
                        headers = {'user-agent': GlobalEnvStorage.UA}
                        r = requests.get(url, params=fetchSuggestion, timeout=10, headers=headers)
                        text = r.text
                        r.close()
                        firstPosition = text.index('[')
                        lastPositon = text.index(']')
                        arrStr = text[firstPosition + 1:lastPositon]
                        arr = arrStr.split(',')
                        str = arr[randint(0, len(arr) - 1)].replace('"', '')
                        str = str.replace('{', '').replace('}', '').replace(':', '').replace('w', '')
    except Exception as e:
        GlobalEnvStorage.infoLogger.info('%s', e)
    finally:
        return str


def fetchRelativeAndRecommendationValue(customerkeyword, terminalType):
    GlobalEnvStorage.infoLogger.info('开始获得相关搜索的词')
    str = None
    try:
        if terminalType == 'PC':
            if GlobalEnvStorage.customerKeyword.searchEngine == '百度':
                url = 'https://www.baidu.com/s'
                payload = {'wd': customerkeyword}
                headers = {'user-agent': GlobalEnvStorage.UA}
                r = requests.get(url, params=payload, timeout=10, headers=headers)
                r.encoding = 'utf-8'
                html = r.text
                r.close()
                page = etree.HTML(html)
                lists = page.xpath('//div[@id="rs"]/table//a')
            else:
                if GlobalEnvStorage.customerKeyword.searchEngine == '搜狗':
                    url = 'https://www.sogou.com/web'
                    payload = {'query': customerkeyword}
                    headers = {'user-agent': GlobalEnvStorage.UA}
                    r = requests.get(url, params=payload, timeout=10, headers=headers)
                    r.encoding = 'utf-8'
                    html = r.text
                    r.close()
                    page = etree.HTML(html)
                    lists = page.xpath('//table[@id="hint_container"]/tbody//a')
                else:
                    if GlobalEnvStorage.customerKeyword.searchEngine == '360':
                        url = 'https://www.so.com/s'
                        payload = {'q': customerkeyword,  'ie': 'utf-8'}
                        r = requests.get(url, params=payload, timeout=10)
                        r.encoding = 'utf-8'
                        html = r.text
                        r.close()
                        page = etree.HTML(html)
                        lists = page.xpath('//div[@id="rs"]/table//a')
        else:
            if GlobalEnvStorage.customerKeyword.searchEngine == '百度':
                url = 'http://m.baidu.com/s'
                payload = {'word': customerkeyword}
                headers = {'user-agent': GlobalEnvStorage.UA}
                r = requests.get(url, params=payload, timeout=10, headers=headers)
                r.encoding = 'utf-8'
                html = r.text
                r.close()
                page = etree.HTML(html)
                lists = page.xpath('//div[@id="relativewords"]/div[@class="rw-list"]/a')
            else:
                if GlobalEnvStorage.customerKeyword.searchEngine == '搜狗':
                    url = 'https://m.sogou.com/web/searchList.jsp'
                    payload = {'keyword': customerkeyword}
                    headers = {'user-agent': GlobalEnvStorage.UA}
                    r = requests.get(url, params=payload, timeout=10, headers=headers)
                    r.encoding = 'utf-8'
                    html = r.text
                    r.close()
                    page = etree.HTML(html)
                    lists = page.xpath('//div[@id="hint"]/ul//a')
                else:
                    if GlobalEnvStorage.customerKeyword.searchEngine == '神马':
                        url = 'http://m.sm.cn/s'
                        payload = {'q': customerkeyword,  'from': 'smor',  'safe': '1'}
                        headers = {'user-agent': GlobalEnvStorage.UA}
                        r = requests.get(url, params=payload, timeout=10, headers=headers)
                        html = r.text
                        r.close()
                        page = etree.HTML(html)
                        lists = page.xpath('//div[@class="sider-card relative-keywords"]//a/@data-log')
                        alist = lists[randint(0, len(lists) - 1)].__str__()
                        arr = eval(alist)
                        arr = arr['q']
        if len(lists) != 0:
            if GlobalEnvStorage.customerKeyword.searchEngine != '神马':
                str = lists[randint(0, len(lists) - 1)].text
            else:
                str = arr
    except Exception as e:
        GlobalEnvStorage.infoLogger.info('%s', e)
    finally:
        return str


def startClear(fisrtRun=False):
    os.system('taskkill /f /im ' + 'chrome.exe')
    os.system('taskkill /f /im ' + 'chromedriver.exe')
    os.system('taskkill /f /im ' + 'rasautou.exe')
    if fisrtRun:
        os.system('taskkill /f /im ' + 'cmd.exe')
        os.system('taskkill /f /im ' + 'chromerefresh.exe')
        os.system('taskkill /f /im ' + 'qq.exe')
        os.system('taskkill /f /im ' + 'notepad.exe')
        os.system('taskkill /f /im ' + 'AutoUpdate.exe')
        os.system('taskkill /f /im ' + 'QQProtect.exe')


def autoStart(name, path):
    KeyName = 'Software\\Microsoft\\Windows\\CurrentVersion\\Run'
    try:
        key = win32api.RegOpenKey(win32con.HKEY_CURRENT_USER, KeyName, 0, win32con.KEY_ALL_ACCESS)
        win32api.RegSetValueEx(key, name, 0, win32con.REG_SZ, path)
        win32api.RegCloseKey(key)
        GlobalEnvStorage.infoLogger.info('添加%s成功！', path)
        unChromerefresh()
    except:
        GlobalEnvStorage.infoLogger.info('error')


def unChromerefresh():
    KeyName = 'Software\\Microsoft\\Windows\\CurrentVersion\\Run'
    try:
        key = win32api.RegOpenKey(win32con.HKEY_LOCAL_MACHINE, KeyName, 0, win32con.KEY_ALL_ACCESS)
        win32api.RegDeleteValue(key, 'chromerefresh')
        win32api.RegCloseKey(key)
        GlobalEnvStorage.infoLogger.info('成功取消旧程序开机启动')
    except:
        GlobalEnvStorage.infoLogger.info('旧程序开机不存在')


def clearLastMain():
    pid = os.getpid()
    ppid = os.getppid()
    psutil.process_iter()
    for proc in psutil.process_iter():
        if proc.name() == 'main.exe' and proc.pid != pid and proc.pid != ppid:
            os.system('taskkill /f /pid ' + str(proc.pid))
            continue


def getExePath():
    path = sys.argv[0]
    GlobalEnvStorage.infoLogger.info('%s', path)
    lastIndex = path.rfind('/')
    parentPath = path[0:lastIndex]
    GlobalEnvStorage.infoLogger.info('%s', parentPath)
    return parentPath


def getLogger(loggname):
    if GlobalEnvStorage.env == 'Development':
        logging.config.fileConfig('C:\\working\\logging.conf')
    else:
        logging.config.fileConfig('../config/dev/logging.conf')
    logger = logging.getLogger(loggname)
    return logger


def KillEventLog(ip, eventName):
    HEvent = win32evtlog.OpenEventLog(ip, eventName)
    if HEvent == None:
        return
    win32evtlog.ClearEventLog(HEvent, None)
    win32evtlog.CloseEventLog(HEvent)


def clearWindowsLog():
    KillEventLog(None, 'Application')
    KillEventLog(None, 'System')
    KillEventLog(None, 'Security')


def profileIDKeywordCount():
    newProfile = False
    if GlobalEnvStorage.profileIDCountList is {} or str(GlobalEnvStorage.profileID) not in GlobalEnvStorage.profileIDCountList:
        profileIDCount = {'count': 1,  'pageSize': 10,  'baiduCookies': False}
        GlobalEnvStorage.profileIDCountList[str(GlobalEnvStorage.profileID)] = profileIDCount
        newProfile = True
    else:
        GlobalEnvStorage.profileIDCountList[str(GlobalEnvStorage.profileID)]['count'] = GlobalEnvStorage.profileIDCountList[str(GlobalEnvStorage.profileID)]['count'] + 1
    return newProfile


def zhangneiUrl(url):
    url = re.sub('\\s', '', url)
    if 'http://' in url:
        url = url.replace('http://', '')
    if 'https://' in url:
        url = url.replace('https://', '')
    if '\\' in url:
        url = url[0:url.find('\\')]
    if '/' in url:
        url = url[0:url.find('/')]
    if '>' in url:
        url = url[0:url.find('>')]
    if '›' in url:
        url = url[0:url.find('›')]
    GlobalEnvStorage.infoLogger.info('%s', url)
    return url


def writeHostFile():
    href = []
    if GlobalEnvStorage.customerKeyword.disableStatistics:
        href.append('hd.baidu.com')
        href.append('hm.baidu.com')
        href.append('tongji.baidu.com')
        href.append('cnzz.com')
        href.append('umeng.com')
        href.append('51.la')
    if GlobalEnvStorage.customerKeyword.disableVisitWebsite:
        href.append(zhangneiUrl(GlobalEnvStorage.customerKeyword.url))
    with open('C:\\WINDOWS\\system32\\drivers\\etc\\hosts', 'w') as (f):
        for h in href:
            f.write('127.0.0.1  ' + h + '\r\n')

        f.write('139.199.84.174 pcsskj.shunshikj.com' + '\r\n')
        f.write('139.199.84.174 msskj.shunshikj.com' + '\r\n')
        f.write('120.24.6.59  dlsskj.shunshikj.com' + '\r\n')
        GlobalEnvStorage.infoLogger.info('加入host的url:%s', href)


def hidewindows():
    hwnd = win32gui.FindWindow('Progman', 'Program Manager')
    hwndSHELLDLL_DefView = win32gui.FindWindowEx(hwnd, None, 'SHELLDLL_DefView', None)
    hwndSysListView32 = win32gui.FindWindowEx(hwndSHELLDLL_DefView, None, 'SysListView32', 'FolderView')
    win32gui.ShowWindow(hwndSysListView32, win32con.SW_HIDE)
    hwnd = win32gui.FindWindow(None, 'C:\\working\\main.exe')
    win32gui.ShowWindow(hwnd, win32con.SW_HIDE)


def logControl():
    if GlobalEnvStorage.loggerState:
        GlobalEnvStorage.infoLogger = getLogger('root')
        GlobalEnvStorage.exceptionlogger = getLogger('root')
        print('日志关闭')
        GlobalEnvStorage.loggerState = False
    else:
        GlobalEnvStorage.infoLogger = getLogger('info')
        GlobalEnvStorage.exceptionlogger = getLogger('error')
        print('日志打开')
        GlobalEnvStorage.loggerState = True


def windowsControl():
    if GlobalEnvStorage.windowsState:
        hwnd = win32gui.FindWindow(None, 'C:\\working\\main.exe')
        win32gui.ShowWindow(hwnd, win32con.SW_HIDE)
        GlobalEnvStorage.windowsState = False
    else:
        hwnd = win32gui.FindWindow(None, 'C:\\working\\main.exe')
        win32gui.ShowWindow(hwnd, win32con.SW_SHOW)
        GlobalEnvStorage.windowsState = True


def hotKeys():
    byref = ctypes.byref
    user32 = ctypes.windll.user32
    HOTKEYS = {1: win32con.VK_F7,
     2: win32con.VK_F6}
    HOTKEY_ACTIONS = {1: logControl,
     2: windowsControl}
    for id, vk in HOTKEYS.items():
        if not user32.RegisterHotKey(None, id, None, vk):
            GlobalEnvStorage.infoLogger.info('Unable to register id，%s', id)
            continue

    msg = wintypes.MSG()
    while user32.GetMessageA(byref(msg), None, 0, 0) != 0:
        if msg.message == win32con.WM_HOTKEY:
            action_to_take = HOTKEY_ACTIONS.get(msg.wParam)
            if action_to_take:
                action_to_take()
            else:
                continue


def clearFile(memoryFreeSpace):
    try:
        if memoryFreeSpace < 500:
            GlobalEnvStorage.infoLogger.info('剩余空间不足500MB,开始清理空间')
            remove_file('C:\\Documents and Settings\\Administrator\\Local Settings\\Temp')
            shutil.rmtree('C:\\working\\cache')
            os.mkdir('C:\\working\\cache')
            shutil.rmtree('C:\\Documents and Settings\\Administrator\\My Documents\\Downloads')
            os.mkdir('C:\\Documents and Settings\\Administrator\\My Documents\\Downloads')
            shutil.rmtree('C:\\Documents and Settings\\Administrator\\Local Settings\\Application Data\\Google')
            os.mkdir('C:\\Documents and Settings\\Administrator\\Local Settings\\Application Data\\Google')
    except Exception as e:
        GlobalEnvStorage.infoLogger.info('%s', e)


def browserQuit():
    try:
        if GlobalEnvStorage.browser:
            GlobalEnvStorage.browser.quit()
    except Exception as e:
        GlobalEnvStorage.infoLogger.info('%s', e)


def remove_file(path, list=None):
    if os.path.isdir(path):
        for childPath in os.listdir(path):
            remove_file(os.path.join(path, childPath), list)

        if os.path.exists(path) and os.listdir(path) == []:
            try:
                os.rmdir(path)
            except:
                pass

    else:
        if os.path.exists(path):
            try:
                os.remove(path)
            except:
                pass


def initCookies():
    if GlobalEnvStorage.cookies == [] and ('_cookies' in GlobalEnvStorage.customerKeyword.operationType or '_pm_map' in GlobalEnvStorage.customerKeyword.operationType):
        try:
            s = open('C:\\working\\cookies.txt', 'r')
            lines = s.readlines()
            for line in lines:
                GlobalEnvStorage.cookies.append(line[:-1])

        finally:
            s.close()


def getUrlByReadExcel(path):
    pass


def getUrlNumber(path):
    pass


def updateReqTimes(path, row, count):
    pass


def cleanReqTimes(path):
    pass

